#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# @Author  : lusheng


import platform
import os
from pylab import *
from openpyxl import load_workbook
import json

def creation_date(path_to_file):
    """
    Try to get the date that a file was created, falling back to when it was
    last modified if that isn't possible.
    See http://stackoverflow.com/a/39501288/1709587 for explanation.
    """
    if platform.system() == 'Windows':
        timestamp = os.path.getmtime(path_to_file)
        timeStruct = time.localtime(timestamp)
        strftime = time.strftime('%Y-%m-%d', timeStruct)
        return strftime
    else:
        stat = os.stat(path_to_file)
        try:
            timestamp = stat.st_birthtime
            timeStruct = time.localtime(timestamp)
            strftime = time.strftime('%Y-%m-%d', timeStruct)
            return strftime
        except AttributeError:
            # We're probably on Linux. No easy way to get creation dates here,
            # so we'll settle for when its content was last modified.
            timestamp = stat.st_mtime
            timeStruct = time.localtime(timestamp)
            strftime = time.strftime('%Y-%m-%d', timeStruct)
            return strftime



def report():
    ptime = creation_date('E:\\python\\webapi\\static\\weekreport\\华诚金属.txt')
    # ptime = creation_date('./static/weekreport/华诚金属.txt')
    excel_path = 'E:\\python\\webapi\\static\\weekreport\\周分析会议数据.xlsx'
    # excel_path = './static/weekreport/周分析会议数据.xlsx'
    wb = load_workbook(excel_path)
    ws = wb.get_sheet_by_name("普氏、MYSTEEL指数")
    print(ws['A%d' % ws.max_row].value)
    date1 = str(ws['A%d' % ws.max_row].value)[:10]
    date1_62 = str(ws['B%d' % ws.max_row].value)
    date1_58 = str(ws['C%d' % ws.max_row].value)
    date2 = str(ws['A%d' % (ws.max_row - 5)].value)[:10]
    date2_62 = str(ws['B%d' % (ws.max_row - 5)].value)
    date2_58 = str(ws['C%d' % (ws.max_row - 5)].value)
    wb.close()
    # print(date1,date2,date1_58,date1_62, date2_58,date2_62)
    if float(date1_62) - float(date2_62) >= 0:
        updown_62 = '上涨'
        diff_62 = round(float(date1_62) - float(date2_62),2)
    else:
        updown_62 = '下跌'
        diff_62 = round(-(float(date1_62) - float(date2_62)),2)

    if float(date1_58) - float(date2_58) >= 0:
        updown_58 = '上涨'
        diff_58 = round(float(date1_58) - float(date2_58),2)
    else:
        updown_58 = '下跌'
        diff_58 = round(-(float(date1_58) - float(date2_58)),2)

    mysteeltext = []
    with open('E:\\python\\webapi\\static\\weekreport\\mysteel.txt', 'r', encoding='utf-8') as f_mysteel:
    # with open('./static/weekreport/mysteel.txt', 'r', encoding='utf-8') as f_mysteel:

        for line in f_mysteel:
            mysteeltext.append(line.strip('\n').split(','))
        # print(mysteeltext)
        kucun_text = mysteeltext[0][0]
        # print(kucun_text)
        kaigong_text = mysteeltext[2][0]
        # print(kaigong_text)
        fenxi_text = mysteeltext[4][0] + mysteeltext[6][0]
        # print(fenxi_text)
        haiyunfei_text = mysteeltext[10][0]
        # print(haiyunfei_text)
        feigang_text = mysteeltext[8][0]
        # print(feigang_text)

    with open('E:\\python\\webapi\\static\\weekreport\\锰矿.txt', 'r', encoding='utf-8') as f_mengkuang:
    # with open('./static/weekreport/锰矿.txt', 'r', encoding='utf-8') as f_mengkuang:
        mengkuang_text = f_mengkuang.read()
    with open('E:\\python\\webapi\\static\\weekreport\\硅锰.txt', 'r', encoding='utf-8') as f_guimeng:
    # with open('./static/weekreport/硅锰.txt', 'r', encoding='utf-8') as f_guimeng:
        guimeng_text = f_guimeng.read()

    stock2 = []
    line = []
    wb = load_workbook('E:\\python\\webapi\\static\\weekreport\\cnfeol1.xlsx')
    # wb = load_workbook('./static/weekreport/cnfeol1.xlsx')
    ws = wb.get_active_sheet()
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            line.append(ws.cell(row=(row + 1), column=(col + 1)).value)
        stock2.append(line)
        line = []
    wb.close()
    # print(stock2)
    menggkuang_kucun = stock2[-1][-2]
    # print(menggkuang_kucun)
    if float(stock2[-1][-2]) - float(stock2[-1][-3]) > 0:
        updown_mengkuang = '增加'
        diff_mengkuang = str(stock2[-1][-1][1:]) + '吨'
    elif float(stock2[-1][-2]) - float(stock2[-1][-3]) < 0:
        updown_mengkuang = '减少'
        diff_mengkuang = str(stock2[-1][-1][1:]) + '吨'
    elif float(stock2[-1][-2]) - float(stock2[-1][-3]) == 0:
        updown_mengkuang = '不变'
        diff_mengkuang = ''

    mengkuang_price = []
    line3 = []
    wb = load_workbook('E:\\python\\webapi\\static\\weekreport\\cnfeol2.xlsx')
    # wb = load_workbook('./static/weekreport/cnfeol2.xlsx')
    ws = wb.get_active_sheet()
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            line3.append(ws.cell(row=(row + 1), column=(col + 1)).value)
        mengkuang_price.append(line3)
        line3 = []
    wb.close()
    # print(mengkuang_price)

    guimeng_price = []
    line4 = []
    wb = load_workbook('E:\\python\\webapi\\static\\weekreport\\cnfeol3.xlsx')
    # wb = load_workbook('./static/weekreport/cnfeol3.xlsx')
    ws = wb.get_active_sheet()
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            line4.append(ws.cell(row=(row + 1), column=(col + 1)).value)
        guimeng_price.append(line4)
        line4 = []
    wb.close()
    # print(guimeng_price)

    with open('E:\\python\\webapi\\static\\weekreport\\华诚金属.txt', 'r', encoding='utf-8') as f_mengpian:
    # with open('./static/weekreport/华诚金属.txt', 'r', encoding='utf-8') as f_mengpian:
        mengpian_text = f_mengpian.read()

    result = {
        'ptime': ptime,
        'date1': date1,
        'date2': date2,
        'date1_62': date1_62,
        'date2_62': date2_62,
        'date1_58': date1_58,
        'date2_58': date2_58,
        'diff_58': diff_58,
        'diff_62': diff_62,
        'updown_58': updown_58,
        'updown_62': updown_62,
        'kucun_text': kucun_text,
        'kaigong_text': kaigong_text,
        'fenxi_text': fenxi_text,
        'haiyunfei_text': haiyunfei_text,
        'feigang_text': feigang_text,
        'mengkuang_text': mengkuang_text,
        'guimeng_text': guimeng_text,
        'stock2': stock2,
        'menggkuang_kucun': menggkuang_kucun,
        'mengkuang_price': mengkuang_price,
        'updown_mengkuang': updown_mengkuang,
        'diff_mengkuang': diff_mengkuang,
        'guimeng_price': guimeng_price,
        'mengpian_text': mengpian_text,
    }
    print(result)
    return result

report()
